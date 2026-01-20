import openpyxl
import sys

def inspect_excel(path):
    print(f"Inspecting: {path}")
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\nSheet: {sheet}")
        print(f"  Dimensions: {ws.dimensions}")
        
        if hasattr(ws, 'tables') and ws.tables:
            print(f"  Tables ({len(ws.tables)}):")
            for tbl_name, tbl_range in ws.tables.items():
                print(f"    - {tbl_name}: {tbl_range}")
        else:
            print("  No Excel Tables found.")

        # Check for non-empty cells to guess layout if no tables
        first_row = ws.min_row
        max_row = min(ws.max_row, 20) # check first 20 rows
        print(f"  First 5 rows content check (min_row={first_row}):")
        for r in range(first_row, first_row + 5):
            vals = [c.value for c in ws[r]]
            # Truncate for display
            display_vals = [str(v)[:10] for v in vals if v is not None]
            if display_vals:
                print(f"    Row {r}: {display_vals}")

if __name__ == "__main__":
    inspect_excel("reports/raw.xlsx")
