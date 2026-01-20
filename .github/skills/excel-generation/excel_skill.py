"""Local copy of the Excel generation utility for the skill.

This is a copy of the workspace utility so the skill folder is self-contained.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import pandas as pd
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def load_input(input_path: str | Path) -> pd.DataFrame:
    p = Path(input_path)
    if str(input_path) == "-":
        raw = sys.stdin.read()
        data = json.loads(raw)
        return pd.DataFrame(data)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {p}")
    if p.suffix.lower() in (".json", ".ndjson"):
        return pd.read_json(p, orient="records")
    if p.suffix.lower() in (".csv", ".tsv"):
        sep = "\t" if p.suffix.lower() == ".tsv" else ","
        return pd.read_csv(p, sep=sep)
    text = p.read_text(encoding="utf-8")
    try:
        data = json.loads(text)
        return pd.DataFrame(data)
    except Exception:
        raise ValueError("Unsupported input format; provide .json or .csv or use stdin JSON")


def load_brand(brand_path: str | Path | None) -> dict[str, Any] | None:
    if not brand_path:
        return None
    p = Path(brand_path)
    if not p.exists():
        print(f"Warning: Brand file not found: {p}", file=sys.stderr)
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        
        # Backward Compatibility: Convert v1 to v2 schema internally
        if "colors" not in data:
            return {
                "name": data.get("name", "Unknown"),
                "fonts": {
                    "heading": data.get("font", "Arial"),
                    "body": data.get("font", "Arial")
                },
                "colors": {
                    "primary": data.get("header_bg", "#ffffff"), # Map header_bg to primary
                    "background": "#ffffff",
                    "text": "#000000",
                    "header_text": data.get("header_font_color", "#000000"),
                    "borders": "#b0aea5" 
                },
                "excel": {
                    "show_gridlines": True, # Default for lazy
                    "header_borders": False
                }
            }
        return data
    except Exception as e:
        print(f"Warning: Failed to parse brand file: {e}", file=sys.stderr)
        return None


def generate_excel(
    df: pd.DataFrame,
    output: str | Path,
    sheet_name: str = "Sheet1",
    engine: str | None = None,
    brand: dict[str, Any] | None = None,
) -> Path:
    out = Path(output)
    out.parent.mkdir(parents=True, exist_ok=True)
    
    # Force openpyxl if branding is requested, as we need it for styling
    if brand and not engine:
        engine = "openpyxl"

    # If not using openpyxl, just write and return
    if engine != "openpyxl" and brand:
        print("Warning: Branding requires 'openpyxl' engine. Styles may not be applied.", file=sys.stderr)

    with pd.ExcelWriter(out, engine=engine or "openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply branding if possible
        if brand and hasattr(writer, "book"):
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            apply_branding(worksheet, df, brand)
            apply_conditional_formatting(worksheet, brand)
            generate_insights(workbook, df, brand)

    return out


def overhaul_excel(input_path: Path, output_path: Path, brand: dict[str, Any]) -> Path:
    """Load an existing Excel file, apply branding, and save to output."""
    wb = load_workbook(input_path)
    
    for sheet_name in wb.sheetnames:
        apply_branding(wb[sheet_name], None, brand)
        apply_conditional_formatting(wb[sheet_name], brand)
    
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def apply_branding(worksheet: Any, df: pd.DataFrame | None, brand: dict[str, Any]) -> None:
    """Apply styles from brand dict to the openpyxl worksheet."""
    
    # Extract brand styles (Assumes v2 normalized structure from load_brand)
    fonts = brand.get("fonts", {})
    colors = brand.get("colors", {})
    opts = brand.get("excel", {})
    
    font_heading = fonts.get("heading", "Arial")
    font_body = fonts.get("body", "Arial")
    
    # Colors (strip hashes)
    c_primary = colors.get("primary", "ffffff").replace("#", "")
    c_header_text = colors.get("header_text", "000000").replace("#", "")
    c_borders = colors.get("borders", "b0aea5").replace("#", "")
    
    # Options
    show_gridlines = opts.get("show_gridlines", True)
    use_borders = opts.get("header_borders", False)
    alternating = opts.get("alternating_rows", False)

    # 0. Sheet Options
    worksheet.sheet_view.showGridLines = show_gridlines

    # 1. Header Styling
    header_font = Font(name=font_heading, bold=True, color=c_header_text)
    header_fill = PatternFill(start_color=c_primary, end_color=c_primary, fill_type="solid")
    
    # Border Style
    thin_border = Side(border_style="thin", color=c_borders)
    bottom_border = Border(bottom=thin_border) if use_borders else Border()

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if use_borders:
            cell.border = bottom_border

    # 2. Body Styling
    body_font = Font(name=font_body)
    
    # Row Iterator (start from 2)
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            
            # Alternating Rows (Zebra Striping)
            if alternating and row_idx % 2 == 0:
                 # Very subtle grey for alternate rows (hardcoded or derived?)
                 # Using a very light grey
                 cell.fill = PatternFill(start_color="f9f9f9", end_color="f9f9f9", fill_type="solid")

    # 3. Auto-size columns
    for i, column in enumerate(worksheet.columns):
        max_length = 0
        column_letter = get_column_letter(i + 1)
        
        # Check header length
        header_val = column[0].value
        if header_val:
            max_length = len(str(header_val))
        
        # Check data length (sample first 100 rows for speed)
        for cell in column[1:101]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.1
        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width

    # 4. Conditional Formatting (Analytics)
    apply_conditional_formatting(worksheet, brand)


def apply_conditional_formatting(worksheet: Any, brand: dict[str, Any]) -> None:
    """Apply rule-based conditional formatting."""
    analytics = brand.get("analytics", {})
    rules = analytics.get("rules", [])
    
    if not rules:
        return

    # Map column names to letters
    # Assumes Row 1 is header
    headers = {}
    for cell in worksheet[1]:
        if cell.value:
            headers[str(cell.value)] = cell.column_letter

    for rule in rules:
        pattern = rule.get("column_pattern")
        condition = rule.get("condition")
        val = rule.get("value")
        style = rule.get("style", {})
        
        if not pattern or not condition:
            continue
            
        # Create DXF Style
        # Note: openpyxl colors for DXF often need ARGB but we allow RGB hex
        dxf_font = None
        dxf_fill = None
        
        if "font_color" in style:
            dxf_font = Font(color=Color(rgb="FF" + style["font_color"].replace("#", "")))
        
        if "bg_color" in style:
            # For DifferentialStyle, PatternFill often needs to be specific
            dxf_fill = PatternFill(start_color=Color(rgb="FF" + style["bg_color"].replace("#", "")), 
                                   end_color=Color(rgb="FF" + style["bg_color"].replace("#", "")),
                                   fill_type="solid")
            
        dxf = DifferentialStyle(font=dxf_font, fill=dxf_fill)

        # Find matching columns
        regex = re.compile(pattern, re.IGNORECASE)
        for col_name, col_letter in headers.items():
            if regex.search(col_name):
                # Apply rule to the whole column (excluding header)
                # E.g. B2:B1048576
                cell_range = f"{col_letter}2:{col_letter}{worksheet.max_row}"
                
                # Construct OpenPyXL Rule
                # Mapping user conditions to OpenPyXL operators
                # lessThan, greaterThan, equal, etc.
                op = condition if condition in ["lessThan", "greaterThan", "equal"] else "lessThan"
                
                formatting_rule = CellIsRule(operator=op, formula=[val], stopIfTrue=True)
                formatting_rule.dxf = dxf
                worksheet.conditional_formatting.add(cell_range, formatting_rule)


def generate_insights(workbook: Any, df: pd.DataFrame, brand: dict[str, Any]) -> None:
    """Auto-generate a chart on a new 'Insights' sheet."""
    # 1. Heuristics to identify Data
    # Identify X-Axis: First text/date column or "ID" column
    x_col = None
    y_cols = []
    
    # Try to find an explicit "X" candidate
    for col in df.columns:
        col_lower = str(col).lower()
        if "date" in col_lower or "time" in col_lower or "id" in col_lower or "batch" in col_lower:
            x_col = col
            break
    
    # If no obvious candidate, pick first object/string column
    if not x_col:
        for col in df.columns:
            if df[col].dtype == 'object':
                x_col = col
                break
                
    # Y-Axes: All numeric columns (excluding X)
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    if x_col and x_col in numeric_cols:
        numeric_cols.remove(x_col)
    
    y_cols = numeric_cols
    
    if not x_col or not y_cols:
         # Cannot chart
         return
         
    # 2. Determine Chart Type
    # If X is Date-like -> LineChart
    # If X is Categorical -> BarChart
    # For now, default to BarChart unless "date"/ "time" in name
    chart_type = BarChart
    if "date" in str(x_col).lower() or "time" in str(x_col).lower():
        chart_type = LineChart
        
    chart = chart_type()
    chart.title = f"{' & '.join(y_cols[:2])} by {x_col}"
    chart.style = 10  # A nice default style
    chart.height = 10
    chart.width = 20
    
    # 3. Define References
    # We need to know which sheet the data is on. Assumes first sheet 'Sheet1' or we find it.
    # In generate_excel context, we know the data sheet name but here we just grab the active one or pass it?
    # We'll assume the FIRST sheet has the data for now.
    data_sheet = workbook.worksheets[0]
    
    # Find column indices (1-based)
    header_row = 1
    # Locate X col index
    x_col_idx = df.columns.get_loc(x_col) + 1
    
    # Cats (X-Axis)
    cats = Reference(data_sheet, min_col=x_col_idx, min_row=header_row+1, max_row=header_row+len(df))
    chart.set_categories(cats)
    
    # Data (Y-Axis)
    # Add first 3 metrics max to avoid clutter
    for i, y_col in enumerate(y_cols[:3]):
        y_col_idx = df.columns.get_loc(y_col) + 1
        data_ref = Reference(data_sheet, min_col=y_col_idx, min_row=header_row, max_row=header_row+len(df))
        series = Series(data_ref, title=str(y_col))
        chart.series.append(series)
        
        # Apply Brand Colors
        # Cycle: primary, secondary, tertiary
        colors = brand.get("colors", {})
        palette = [
            colors.get("primary", "d97757").replace("#", ""),
            colors.get("secondary", "6a9bcc").replace("#", ""),
            colors.get("tertiary", "788c5d").replace("#", "")
        ]
        color_hex = palette[i % len(palette)]
        
        # OpenPyXL Chart coloring
        # For BarChart it's graphicalProperties.solidFill
        # For LineChart it's graphicalProperties.line.solidFill
        # We'll try generic implementation
        if hasattr(series.graphicalProperties, "solidFill"):
             series.graphicalProperties.solidFill = color_hex
        if hasattr(series.graphicalProperties, "line"):
             series.graphicalProperties.line.solidFill = color_hex

    # 4. output
    # Create Insights sheet
    if "Insights" in workbook.sheetnames:
        ws_chart = workbook["Insights"]
    else:
        ws_chart = workbook.create_sheet("Insights")
        
    ws_chart.add_chart(chart, "B2")



def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate .xlsx from JSON/CSV/stdin")
    parser.add_argument("--input", "-i", required=True, help="Path to input file or '-' for stdin (JSON)")
    parser.add_argument("--output", "-o", required=True, help="Output .xlsx path")
    parser.add_argument("--sheet", "-s", default="Sheet1", help="Sheet name")
    parser.add_argument("--engine", "-e", default=None, help="Optional pandas Excel engine (xlsxwriter, openpyxl)")
    parser.add_argument("--brand", "-b", default=None, help="Path to brand JSON file")
    args = parser.parse_args(argv)


    # Special handling for re-branding existing Excel files
    input_path = Path(args.input)
    if input_path.suffix.lower() == ".xlsx" and input_path.exists():
        if not args.brand:
            print("Error: --brand is required when input is an .xlsx file (re-branding mode).", file=sys.stderr)
            return 1
        
        try:
            brand_data = load_brand(args.brand)
            out = overhaul_excel(input_path, args.output, brand_data)
            print(str(out))
            return 0
        except Exception as ex:
            print(f"Failed to overhaul Excel: {ex}", file=sys.stderr)
            return 3

    try:
        df = load_input(args.input)
    except Exception as ex:
        print(f"Failed to load input: {ex}", file=sys.stderr)
        return 2


    try:
        brand_data = load_brand(args.brand)
        out = generate_excel(df, args.output, sheet_name=args.sheet, engine=args.engine, brand=brand_data)
        print(str(out))
        return 0
    except Exception as ex:
        print(f"Failed to write Excel: {ex}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
